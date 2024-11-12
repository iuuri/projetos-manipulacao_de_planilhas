''' 
workbook = Planilha
sheet = página
'''
import openpyxl
# Criação de workbook
workbook = openpyxl.Workbook()

# Mostrar sheets existentes
print(workbook.sheetnames)

# Para criar novos sheets
workbook.create_sheet('ruas')
workbook.create_sheet('cidades')
workbook.create_sheet('estados')
print(workbook.sheetnames)

# Salvar modificações para uma planilha
workbook.save('endereços.xlsx')

# Alterar o nome de um sheet
workbook['ruas'].title = 'ruas da cidade'
workbook.save('endereços.xlsx')

# Como excluir um sheet
del workbook['Sheet']
print(workbook.sheetnames)
workbook.save('endereços.xlsx')