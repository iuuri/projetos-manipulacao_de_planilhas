import openpyxl

workbook = openpyxl.load_workbook('hockey_players.xlsx')
print(workbook.sheetnames)
pagina = input('Qual página gostaria de ler?: ')
sheet_usuario = workbook[pagina]
linha_minima = int(input('Em qual linha inicia a leitura?'))
linha_maxima = int(input('Em qual linha gostaria de finalizar a leitura?'))
coluna_minima = int(input('Em qual coluna devo iniciar?: '))
coluna_maxima = int(
    input('Qual é a coluna onde devemos finalizar a pesquisa: '))

for linha in sheet_usuario.iter_rows(min_row=linha_minima, max_row=linha_maxima, min_col=coluna_minima, max_col=coluna_maxima):
    print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, linha[4].value,
          linha[5].value, linha[6].value, linha[7].value, linha[8].value, linha[9].value, linha[10].value)